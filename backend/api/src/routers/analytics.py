from __future__ import annotations

from datetime import datetime, timedelta
from typing import Literal

from fastapi import APIRouter, Depends
from sqlalchemy.ext.asyncio import AsyncSession

from ..balance_sheet import generate_pnl, generate_balance_sheet, generate_cash_flow
from ..cache.analytics_cache import analytics_cache
from ..database import get_db
from ..repositories.transaction import TransactionRepo
from ..security import AuthContext, require_plan

router = APIRouter(prefix="/analytics", tags=["analytics"])

_RANGES: dict[str, timedelta] = {
    "1m": timedelta(days=30),
    "3m": timedelta(days=90),
    "6m": timedelta(days=180),
    "1y": timedelta(days=365),
}


@router.get("/pnl", response_model=dict)
async def analytics_pnl(
    range: Literal["1y", "6m", "3m", "1m"] = "3m",
    auth: AuthContext = Depends(require_plan("analytics_basic")),
    session: AsyncSession = Depends(get_db),
):
    now = datetime.utcnow()
    window = _RANGES[range]
    start = now - window

    cached = await analytics_cache.get_pnl(auth.user.org_id, range)
    if cached:
        return cached

    trepo = TransactionRepo()
    tx = await trepo.window(session, auth.user.org_id, start, now)
    
    # Also fetch Journal Entries
    from sqlalchemy import select
    from ..models import JournalEntry, JournalDay
    
    journal_stmt = (
        select(JournalEntry, JournalDay.journal_date)
        .join(JournalDay, JournalEntry.journal_day_id == JournalDay.id)
        .where(
            JournalEntry.org_id == auth.user.org_id,
            JournalDay.journal_date >= start.date(),
            JournalDay.journal_date <= now.date(),
            # Only include valid entries
            JournalEntry.entry_type.in_(["revenue", "cost", "inventory_purchase", "inventory_use", "transfer"])
        )
    )
    journal_res = await session.execute(journal_stmt)
    journal_rows = journal_res.all()

    # Normalize Transactions
    rows = [
        {
            "Account": t.account_code,
            "Category": t.category,
            "Amount": float(t.amount),
            "Date": t.txn_date,
            "Description": t.description,
        }
        for t in tx
    ]
    
    # Normalize Journal Entries
    for entry, date_val in journal_rows:
        amount = float(entry.total)
        # Flip sign for expenses
        if entry.entry_type in ("cost", "inventory_purchase"):
            amount = -abs(amount)
        
        rows.append({
            "Account": entry.item_name or "Journal Entry",
            "Category": entry.category or "Uncategorized",
            "Amount": amount,
            "Date": datetime(date_val.year, date_val.month, date_val.day),
            "Description": entry.notes or entry.item_name,
        })

    pnl = generate_pnl(rows)
    expenses_total = pnl["cogs"] + pnl["total_expenses"]

    series_map: dict[tuple[int, int, int], dict[str, float | str]] = {}
    
    # Process combined rows for series - use daily granularity
    for r in rows:
        d = r["Date"]
        # Handle cases where d might be date or datetime
        if hasattr(d, "date"): 
             # It's a datetime
             period_key = (d.year, d.month, d.day)
             iso_date = d.date().isoformat()
        else:
             # It's a date
             period_key = (d.year, d.month, d.day)
             iso_date = d.isoformat()

        entry = series_map.setdefault(
            period_key,
            {
                "date": iso_date,
                "revenue": 0.0,
                "expenses": 0.0,
            },
        )
        amount = r["Amount"]
        if amount >= 0:
            entry["revenue"] = float(entry["revenue"]) + amount
        else:
            entry["expenses"] = float(entry["expenses"]) + abs(amount)

    series = [series_map[key] for key in sorted(series_map.keys())]

    result = {
        "range": range,
        "revenue": pnl["revenue"],
        "expenses": expenses_total,
        "profit": pnl["net_income"],
        "series": series,
    }
    await analytics_cache.set_pnl(auth.user.org_id, range, result)
    return result


@router.get("/receivables", response_model=dict)
async def analytics_receivables(
    auth: AuthContext = Depends(require_plan("analytics_basic")),
    session: AsyncSession = Depends(get_db),
):
    """Get accounts receivable summary - unpaid revenue entries (money owed TO user)."""
    from sqlalchemy import select, func
    from ..models import JournalEntry, Transaction
    
    # Get unpaid revenue from journal entries
    journal_query = select(
        JournalEntry.category,
        func.sum(JournalEntry.total).label("total"),
        func.count(JournalEntry.id).label("count"),
    ).where(
        JournalEntry.org_id == auth.user.org_id,
        JournalEntry.payment_status == "unpaid",
        JournalEntry.entry_type == "revenue",
    ).group_by(JournalEntry.category)
    
    journal_result = await session.execute(journal_query)
    journal_rows = journal_result.all()
    
    # Get unpaid revenue from transactions (positive amounts = revenue)
    txn_query = select(
        Transaction.category,
        func.sum(Transaction.amount).label("total"),
        func.count(Transaction.id).label("count"),
    ).where(
        Transaction.org_id == auth.user.org_id,
        Transaction.payment_status == "unpaid",
        Transaction.amount > 0,
    ).group_by(Transaction.category)
    
    txn_result = await session.execute(txn_query)
    txn_rows = txn_result.all()
    
    # Combine results
    by_category: dict[str, dict] = {}
    total_receivable = 0.0
    total_count = 0
    
    for row in journal_rows:
        cat = row.category or "Uncategorized"
        amount = float(row.total or 0)
        by_category[cat] = {"amount": amount, "count": row.count}
        total_receivable += amount
        total_count += row.count
    
    for row in txn_rows:
        cat = row.category or "Uncategorized"
        amount = float(row.total or 0)
        if cat in by_category:
            by_category[cat]["amount"] += amount
            by_category[cat]["count"] += row.count
        else:
            by_category[cat] = {"amount": amount, "count": row.count}
        total_receivable += amount
        total_count += row.count
    
    breakdown = [
        {"category": k, "amount": v["amount"], "count": v["count"]}
        for k, v in sorted(by_category.items(), key=lambda x: -x[1]["amount"])
    ]
    
    return {
        "total": total_receivable,
        "count": total_count,
        "breakdown": breakdown,
    }


@router.get("/payables", response_model=dict)
async def analytics_payables(
    auth: AuthContext = Depends(require_plan("analytics_basic")),
    session: AsyncSession = Depends(get_db),
):
    """Get accounts payable summary - unpaid expense entries (money owed BY user)."""
    from sqlalchemy import select, func
    from ..models import JournalEntry, Transaction
    
    # Get unpaid expenses from journal entries (cost type)
    journal_query = select(
        JournalEntry.category,
        func.sum(JournalEntry.total).label("total"),
        func.count(JournalEntry.id).label("count"),
    ).where(
        JournalEntry.org_id == auth.user.org_id,
        JournalEntry.payment_status == "unpaid",
        JournalEntry.entry_type.in_(["cost", "inventory_purchase"]),
    ).group_by(JournalEntry.category)
    
    journal_result = await session.execute(journal_query)
    journal_rows = journal_result.all()
    
    # Get unpaid expenses from transactions (negative amounts = expenses)
    txn_query = select(
        Transaction.category,
        func.sum(func.abs(Transaction.amount)).label("total"),
        func.count(Transaction.id).label("count"),
    ).where(
        Transaction.org_id == auth.user.org_id,
        Transaction.payment_status == "unpaid",
        Transaction.amount < 0,
    ).group_by(Transaction.category)
    
    txn_result = await session.execute(txn_query)
    txn_rows = txn_result.all()
    
    # Combine results
    by_category: dict[str, dict] = {}
    total_payable = 0.0
    total_count = 0
    
    for row in journal_rows:
        cat = row.category or "Uncategorized"
        amount = float(row.total or 0)
        by_category[cat] = {"amount": amount, "count": row.count}
        total_payable += amount
        total_count += row.count
    
    for row in txn_rows:
        cat = row.category or "Uncategorized"
        amount = float(row.total or 0)
        if cat in by_category:
            by_category[cat]["amount"] += amount
            by_category[cat]["count"] += row.count
        else:
            by_category[cat] = {"amount": amount, "count": row.count}
        total_payable += amount
        total_count += row.count
    
    breakdown = [
        {"category": k, "amount": v["amount"], "count": v["count"]}
        for k, v in sorted(by_category.items(), key=lambda x: -x[1]["amount"])
    ]
    
    return {
        "total": total_payable,
        "count": total_count,
        "breakdown": breakdown,
    }


@router.get("/receivables/list", response_model=list)
async def list_receivables(
    auth: AuthContext = Depends(require_plan("analytics_basic")),
    session: AsyncSession = Depends(get_db),
):
    """List individual unpaid revenue transactions for drill-down view."""
    from sqlalchemy import select, union_all
    from ..models import JournalEntry, JournalDay, Transaction
    
    # Get unpaid revenue from journal entries
    journal_query = (
        select(
            JournalEntry.id.label("id"),
            JournalEntry.item_name.label("description"),
            JournalEntry.total.label("amount"),
            JournalEntry.category,
            JournalEntry.created_at.label("date"),
            JournalEntry.payment_status,
        )
        .select_from(JournalEntry)
        .where(
            JournalEntry.org_id == auth.user.org_id,
            JournalEntry.payment_status == "unpaid",
            JournalEntry.entry_type == "revenue",
        )
    )
    
    journal_result = await session.execute(journal_query)
    
    # Get unpaid revenue from transactions
    txn_query = select(
        Transaction.id,
        Transaction.description,
        Transaction.amount,
        Transaction.category,
        Transaction.txn_date.label("date"),
        Transaction.payment_status,
    ).where(
        Transaction.org_id == auth.user.org_id,
        Transaction.payment_status == "unpaid",
        Transaction.amount > 0,
    )
    
    txn_result = await session.execute(txn_query)
    
    items = []
    for row in journal_result.all():
        items.append({
            "id": row.id,
            "type": "journal",
            "description": row.description or "Unnamed",
            "amount": float(row.amount or 0),
            "category": row.category or "Uncategorized",
            "date": row.date.isoformat() if row.date else None,
        })
    
    for row in txn_result.all():
        items.append({
            "id": row.id,
            "type": "transaction",
            "description": row.description or "Unnamed",
            "amount": float(row.amount or 0),
            "category": row.category or "Uncategorized",
            "date": row.date.isoformat() if row.date else None,
        })
    
    return sorted(items, key=lambda x: x["amount"], reverse=True)


@router.get("/payables/list", response_model=list)
async def list_payables(
    auth: AuthContext = Depends(require_plan("analytics_basic")),
    session: AsyncSession = Depends(get_db),
):
    """List individual unpaid expense transactions for drill-down view."""
    from sqlalchemy import select, func
    from ..models import JournalEntry, Transaction
    
    # Get unpaid expenses from journal entries
    journal_query = select(
        JournalEntry.id,
        JournalEntry.item_name.label("description"),
        JournalEntry.total.label("amount"),
        JournalEntry.category,
        JournalEntry.created_at.label("date"),
    ).where(
        JournalEntry.org_id == auth.user.org_id,
        JournalEntry.payment_status == "unpaid",
        JournalEntry.entry_type.in_(["cost", "inventory_purchase"]),
    )
    
    journal_result = await session.execute(journal_query)
    
    # Get unpaid expenses from transactions
    txn_query = select(
        Transaction.id,
        Transaction.description,
        func.abs(Transaction.amount).label("amount"),
        Transaction.category,
        Transaction.txn_date.label("date"),
    ).where(
        Transaction.org_id == auth.user.org_id,
        Transaction.payment_status == "unpaid",
        Transaction.amount < 0,
    )
    
    txn_result = await session.execute(txn_query)
    
    items = []
    for row in journal_result.all():
        items.append({
            "id": row.id,
            "type": "journal",
            "description": row.description or "Unnamed",
            "amount": float(row.amount or 0),
            "category": row.category or "Uncategorized",
            "date": row.date.isoformat() if row.date else None,
        })
    
    for row in txn_result.all():
        items.append({
            "id": row.id,
            "type": "transaction",
            "description": row.description or "Unnamed",
            "amount": float(row.amount or 0),
            "category": row.category or "Uncategorized",
            "date": row.date.isoformat() if row.date else None,
        })
    
    return sorted(items, key=lambda x: x["amount"], reverse=True)


@router.patch("/transaction/{txn_id}/payment-status")
async def toggle_transaction_payment_status(
    txn_id: int,
    status: str,  # "paid" | "unpaid"
    auth: AuthContext = Depends(require_plan("analytics_basic")),
    session: AsyncSession = Depends(get_db),
):
    """Toggle payment status for a transaction (from Excel uploads). Used for AR/AP tracking."""
    from sqlalchemy import select
    from datetime import datetime
    from ..models import Transaction
    
    if status not in ("paid", "unpaid"):
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="status must be 'paid' or 'unpaid'")
    
    # Verify transaction exists and belongs to user's org
    query = select(Transaction).where(
        Transaction.id == txn_id,
        Transaction.org_id == auth.user.org_id,
    )
    result = await session.execute(query)
    txn = result.scalar_one_or_none()
    
    if not txn:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    # Update payment status
    payment_date = datetime.utcnow() if status == "paid" else None
    txn.payment_status = status
    txn.payment_date = payment_date
    
    await session.commit()
    
    return {
        "id": txn.id,
        "payment_status": status,
        "payment_date": payment_date.isoformat() if payment_date else None,
    }


@router.post("/generate")
async def generate_financial_document(
    doc_type: Literal["balance_sheet", "pnl", "cash_flow"],
    range: Literal["1y", "6m", "3m", "1m", "all"] = "3m",
    auth: AuthContext = Depends(require_plan("analytics_basic")),
    session: AsyncSession = Depends(get_db),
):
    """
    Generate a financial document as a downloadable Excel file.
    
    Supports:
    - balance_sheet: Assets, Liabilities, Equity
    - pnl: Profit & Loss Statement (Income Statement)
    - cash_flow: Cash Flow Statement (Operating, Investing, Financing)
    """
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from sqlalchemy import select
    from ..models import Transaction, JournalEntry, JournalDay
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    # Calculate date range
    now = datetime.utcnow()
    if range == "all":
        start = datetime(2000, 1, 1)
    else:
        window = _RANGES.get(range, timedelta(days=90))
        start = now - window
    
    # Fetch transactions
    trepo = TransactionRepo()
    transactions = await trepo.window(session, auth.user.org_id, start, now)
    
    # Also fetch Journal Entries
    journal_stmt = (
        select(JournalEntry, JournalDay.journal_date)
        .join(JournalDay, JournalEntry.journal_day_id == JournalDay.id)
        .where(
            JournalEntry.org_id == auth.user.org_id,
            JournalDay.journal_date >= start.date(),
            JournalDay.journal_date <= now.date(),
        )
    )
    journal_res = await session.execute(journal_stmt)
    journal_rows = journal_res.all()
    
    # Normalize data
    rows = []
    for t in transactions:
        rows.append({
            "Account": t.account_code,
            "Category": t.category,
            "Amount": float(t.amount),
            "Date": t.txn_date,
            "Description": t.description,
        })
    
    for entry, date_val in journal_rows:
        amount = float(entry.total)
        if entry.entry_type in ("cost", "inventory_purchase"):
            amount = -abs(amount)
        rows.append({
            "Account": entry.item_name or "Journal Entry",
            "Category": entry.category or entry.entry_type,
            "Amount": amount,
            "Date": date_val,
            "Description": entry.item_name,
        })
    
    # Generate report data
    if doc_type == "balance_sheet":
        result = generate_balance_sheet(rows)
        doc_title = "Balance Sheet"
    elif doc_type == "pnl":
        result = generate_pnl(rows)
        doc_title = "Profit & Loss Statement"
    elif doc_type == "cash_flow":
        result = generate_cash_flow(rows)
        doc_title = "Cash Flow Statement"
    else:
        result = {}
        doc_title = "Financial Report"
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = doc_title[:31]  # Excel limits sheet names to 31 chars
    
    # Styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    money_format = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row_num = 1
    
    # Title
    ws.cell(row=row_num, column=1, value=doc_title)
    ws.cell(row=row_num, column=1).font = Font(bold=True, size=16)
    row_num += 1
    
    # Period
    ws.cell(row=row_num, column=1, value=f"Period: {start.date().isoformat()} to {now.date().isoformat()}")
    ws.cell(row=row_num, column=1).font = Font(italic=True, size=10)
    row_num += 2
    
    # Generate content based on document type
    if doc_type == "balance_sheet":
        for section in ["assets", "liabilities", "equity"]:
            if section in result and result[section]:
                # Section header
                ws.cell(row=row_num, column=1, value=section.upper())
                ws.cell(row=row_num, column=1).font = header_font
                ws.cell(row=row_num, column=1).fill = header_fill
                ws.cell(row=row_num, column=2).fill = header_fill
                row_num += 1
                
                # Items
                for account, amount in result[section].items():
                    ws.cell(row=row_num, column=1, value=account)
                    ws.cell(row=row_num, column=2, value=amount)
                    ws.cell(row=row_num, column=2).number_format = money_format
                    row_num += 1
                
                row_num += 1
        
        # Totals
        if "totals" in result:
            ws.cell(row=row_num, column=1, value="TOTALS")
            ws.cell(row=row_num, column=1).font = subheader_font
            row_num += 1
            for key, value in result["totals"].items():
                ws.cell(row=row_num, column=1, value=key.replace("_", " ").title())
                ws.cell(row=row_num, column=2, value=value)
                ws.cell(row=row_num, column=2).number_format = money_format
                row_num += 1
    
    elif doc_type == "pnl":
        # Revenue
        ws.cell(row=row_num, column=1, value="Revenue")
        ws.cell(row=row_num, column=1).font = subheader_font
        ws.cell(row=row_num, column=2, value=result.get("revenue", 0))
        ws.cell(row=row_num, column=2).number_format = money_format
        row_num += 1
        
        # COGS
        ws.cell(row=row_num, column=1, value="Cost of Goods Sold")
        ws.cell(row=row_num, column=2, value=result.get("cogs", 0))
        ws.cell(row=row_num, column=2).number_format = money_format
        row_num += 1
        
        # Gross Profit
        ws.cell(row=row_num, column=1, value="Gross Profit")
        ws.cell(row=row_num, column=1).font = subheader_font
        ws.cell(row=row_num, column=2, value=result.get("gross_profit", 0))
        ws.cell(row=row_num, column=2).number_format = money_format
        row_num += 2
        
        # Operating Expenses
        ws.cell(row=row_num, column=1, value="OPERATING EXPENSES")
        ws.cell(row=row_num, column=1).font = header_font
        ws.cell(row=row_num, column=1).fill = header_fill
        ws.cell(row=row_num, column=2).fill = header_fill
        row_num += 1
        
        for expense, amount in result.get("expenses", {}).items():
            ws.cell(row=row_num, column=1, value=expense)
            ws.cell(row=row_num, column=2, value=amount)
            ws.cell(row=row_num, column=2).number_format = money_format
            row_num += 1
        
        row_num += 1
        ws.cell(row=row_num, column=1, value="Total Expenses")
        ws.cell(row=row_num, column=1).font = subheader_font
        ws.cell(row=row_num, column=2, value=result.get("total_expenses", 0))
        ws.cell(row=row_num, column=2).number_format = money_format
        row_num += 2
        
        # Net Income
        ws.cell(row=row_num, column=1, value="NET INCOME")
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row_num, column=2, value=result.get("net_income", 0))
        ws.cell(row=row_num, column=2).number_format = money_format
        ws.cell(row=row_num, column=2).font = Font(bold=True)
    
    elif doc_type == "cash_flow":
        for activity in ["operating_activities", "investing_activities", "financing_activities"]:
            if activity in result:
                section = result[activity]
                # Section header
                ws.cell(row=row_num, column=1, value=activity.replace("_", " ").upper())
                ws.cell(row=row_num, column=1).font = header_font
                ws.cell(row=row_num, column=1).fill = header_fill
                ws.cell(row=row_num, column=2).fill = header_fill
                row_num += 1
                
                ws.cell(row=row_num, column=1, value="Inflows")
                ws.cell(row=row_num, column=2, value=section.get("inflows", 0))
                ws.cell(row=row_num, column=2).number_format = money_format
                row_num += 1
                
                ws.cell(row=row_num, column=1, value="Outflows")
                ws.cell(row=row_num, column=2, value=section.get("outflows", 0))
                ws.cell(row=row_num, column=2).number_format = money_format
                row_num += 1
                
                ws.cell(row=row_num, column=1, value="Net")
                ws.cell(row=row_num, column=1).font = subheader_font
                ws.cell(row=row_num, column=2, value=section.get("net", 0))
                ws.cell(row=row_num, column=2).number_format = money_format
                row_num += 2
        
        # Net change
        ws.cell(row=row_num, column=1, value="NET CHANGE IN CASH")
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row_num, column=2, value=result.get("net_change_in_cash", 0))
        ws.cell(row=row_num, column=2).number_format = money_format
        ws.cell(row=row_num, column=2).font = Font(bold=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Generate filename
    filename = f"{doc_type}_{start.date().isoformat()}_to_{now.date().isoformat()}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
